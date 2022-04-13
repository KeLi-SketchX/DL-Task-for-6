import numpy as np
import pandas as pd
import openpyxl
import os
from sklearn.metrics import accuracy_score,classification_report,confusion_matrix
from openpyxl import Workbook, load_workbook
import time

def write2excel(img_names, pred_labels):
    for i in range(len(img_names)):
        img_names[i] = img_names[i].split('/')[-1].split('.')[0]
    if len(img_names) != len(pred_labels):
        print('图片名列表与结果列表长度不同，请检查！')
        return
    a = 'Sketch_Recognition_Challenge_'+time.strftime('%Y_%m_%d_%H_%M_%S') + '.xlsx'
    if os.path.exists(a) == True:
        print('当前文件夹下有同名文件，请删除或等待一秒后重试！')
        return
    else:
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(title = 'sheet1', index = 0)
        data = []
        list1 = img_names
        list2 = pred_labels
        for i in range(0,len(img_names)):
            data.append([list1[i],list2[i]])
        for i in data:
            sheet.append(i)
        wb.save(a)
        print('生成成功！生成的文件名字为：' + a)
        return a

def write_text(filename,string):
    file_path = filename
    file = open(file_path,'w')
    file.write(string)
    file.close()

def evaluate(excelpath,seen=1):
    # --------------------------------------------------------------------
    # 将test_unseen_label.txt的数据存为字典
    rate1 = open('test_unseen_label.txt', 'r', encoding='utf-8')
    test_unseen_label = dict()
    for line in rate1:
        line = line.strip().split(' ')
        test_unseen_label[line[0]] = line[1]
    rate1.close()
    # 将test_seen_label.txt的数据存为字典
    rate1 = open('test_seen_label.txt', 'r', encoding='utf-8')
    test_seen_label = dict()
    for line in rate1:
        line = line.strip().split(' ')
        test_seen_label[line[0]] = line[1]
    rate1.close()
    # --------------------------------------------------------------------
    file_path = excelpath
    wb = load_workbook(file_path)
    sheet_list = wb.sheetnames
    ws = wb[sheet_list[0]]
    total_list = []
    for row in ws.rows:
        row_list = []
        row_list.append(str(row[0].value))
        row_list.append(str(row[1].value))
        total_list.append(row_list)
    pred=dict(total_list)
    # --------------------------------------------------------------------
    y_true = []
    y_pred = []
    x = pred
    #seen =1
    if seen==0 :
        y = test_unseen_label
    else:
        y= test_seen_label
    z = y.keys()-y.keys()^x.keys()
    for key in z:
        for i,j in y.items():
            if i==key:
                y_pred.append(x[key])
                y_true.append(j)
    str1='------------------测试集上得分：---------------------—-'
    str2='------------------- 混淆矩阵: ----------------------—-'
    str3='矩阵类别序列同上'
    str4=str1+'\n'+classification_report(y_true, y_pred)+'\n'+str2+'\n'+str3+'\n'
    print(str4)

    output_file = 'evaluation_'+time.strftime('%Y_%m_%d_%H_%M_%S')+'.txt'
    write_text(output_file,str4)

    with open(output_file, 'a') as f:
        np.savetxt(f, np.column_stack(confusion_matrix(y_true, y_pred).T), fmt='%10.f')
    # confusion_matrix(y_true, y_pred)
    print(confusion_matrix(y_true, y_pred))

if __name__=="__main__":
    img_names = ['19142', '19143', '19144', '19145']
    pred_labels = ['mouse', 'backpack', 'backpack', 'backpack']
    b = write2excel(img_names, pred_labels)
    evaluate(b)